"""
Build the spreadsheet that other people fill in to add listings.

# Why a spreadsheet and not a small app

The people filling this in are not on the team and will not be trained. A form
would need building, hosting, an account each, and a way to save half-finished
work; a spreadsheet is a thing they already know, works offline, and can be
emailed. What a form would have bought is validation, and Excel's own dropdowns
buy most of that at the point where it matters - while they are typing, rather
than in a report afterwards.

The columns are not a design. They are exactly what import/listing-row.ts reads,
because a template that produces a file the importer cannot read is worse than
no template. Anything the importer ignores is marked as such rather than quietly
included, so nobody spends an afternoon filling a column that goes nowhere.

# What this cannot express

The importer hardcodes Mount Lebanon and knows three district headings. A
listing anywhere else needs the importer changed first, so the district list
here is short on purpose rather than by oversight.

Run: python scripts/make-listing-template.py
"""

import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = "vardenia-listing-template.xlsx"
HERE = os.path.dirname(os.path.abspath(__file__))

# Straight from CATEGORY_BY_HEADING in import/listing-row.ts. Anything not on
# this list is reported as "unknown category" and the row is not imported.
CATEGORIES = [
    "Hotels",
    "Guest Houses",
    "Restaurants",
    "Activities",
    "Tour Guides",
    "Festivals",
]

# From DISTRICT_BY_HEADING. Short because the importer only covers Mount Lebanon.
DISTRICTS = [
    "Keserwan District",
    "Byblos / Jbeil District",
    "Keserwan + Byblos / Jbeil Districts",
]

# seasonalityFrom() looks for these words. Anything else contributes nothing.
SEASONS = ["Year-round", "Summer", "Winter", "Summer and Winter"]

STARS = ["1", "2", "3", "4", "5"]

# These four strings are not labels. priceBand() reads the first number after
# the dollar sign, so they land in bands 1, 2, 3 and 4 exactly - verified, not
# assumed. Changing the wording without re-checking would silently re-band
# every listing that uses it.
PRICES = ["$0-49", "$50-149", "$150-299", "$300+"]

# The towns already in the directory, cleaned. Free typing produced "FARAYA",
# "Faraya" and "Faraya (Airbnb)" as three different places in the last import,
# which is what this list exists to stop.
with open(os.path.join(HERE, "towns.json"), encoding="utf-8") as handle:
    TOWNS = json.load(handle)

# (header, width, help text, dropdown list or None, required, strict)
#
# `strict` is the difference between a list that refuses anything else and one
# that only warns. A category outside the list cannot be imported, so it
# refuses. A town outside the list is simply a town nobody has entered yet, so
# it warns and lets them continue - a locked list would make the sheet unusable
# the first time a new village appears.
COLUMNS = [
    ("Name / Listing", 34, "The business name as it should appear on the site. Nothing else - no town, no phone.", None, True, True),
    ("Category", 20, "Pick from the list. This decides which section of the site the listing appears in.", CATEGORIES, True, True),
    ("District", 26, "Pick from the list.", DISTRICTS, True, True),
    ("Location", 26, "Town or village. Pick from the list; type a new one only if it is genuinely not there.", TOWNS, False, False),
    ("Type / Activity", 30, "What it is, in a few words. Becomes the listing's tags. Separate several with commas.", None, False, True),
    ("Hotel Stars", 11, "Hotels only. Pick 1 to 5. Under 4 files the hotel as boutique rather than luxury.", STARS, False, True),
    ("Price Range", 16, "Pick the band a typical visit or night costs.", PRICES, False, False),
    ("Rating / 5", 11, "The Google rating, 0 to 5. Leave empty if you have not checked it.", None, False, True),
    ("Usually When", 20, "When it is open or worth visiting.", SEASONS, False, True),
    ("Overview / Description", 60, "A short paragraph. The first sentence becomes the tagline under the name.", None, False, True),
]

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HELP_FILL = PatternFill("solid", fgColor="F3F4F6")
REQUIRED_FILL = PatternFill("solid", fgColor="FEF3C7")

book = Workbook()

# --- instructions ----------------------------------------------------------
guide = book.active
guide.title = "Read me first"
guide.column_dimensions["A"].width = 100

GUIDE = [
    ("Vardenia - adding listings", "title"),
    ("", None),
    ("Fill in the Listings tab. One business per row. Do not rename or reorder the columns:", None),
    ("the importer finds them by name and will skip anything it does not recognise.", None),
    ("", None),
    ("The two orange columns are required. A row without them cannot be imported.", None),
    ("Everything else can be left empty and added later.", None),
    ("", None),
    ("Most columns are dropdowns - click the cell and a little arrow appears. Use them.", None),
    ("Only the business name, the activity and the description have to be typed.", None),
    ("", None),
    ("Category, District, Hotel Stars and Usually When will refuse anything not on the list,", None),
    ("because a value off the list cannot be imported at all.", None),
    ("", None),
    ("Location and Price Range only warn. If a village is genuinely not on the list, type it", None),
    ("and carry on. Otherwise pick the one that is there - Faraya typed three different ways", None),
    ("becomes three different places on the site.", None),
    ("", None),
    ("Only Mount Lebanon for now", "head"),
    ("The importer currently covers Keserwan and Byblos / Jbeil only. A business anywhere", None),
    ("else needs a change to the site first - send it separately rather than inventing a district.", None),
    ("", None),
    ("Photos", "head"),
    ("Do not put photos in this file. Photos go in folders, one folder per business.", None),
    ("Send this sheet back first. We then send you a folder tree - one empty folder per", None),
    ("business, already named, with a note inside saying what goes in it. Drop the photos", None),
    ("into the folders and send the whole thing back.", None),
    ("", None),
    ("Do not rename the folders. The name is how each photo finds its listing.", None),
    ("", None),
    ("Inside each folder:", None),
    ("    cover.jpg    the main photo, the one that appears at the top of the listing", None),
    ("    01.jpg       gallery photos, in the order you want them shown", None),
    ("    02.jpg", None),
    ("", None),
    ("Send cover plus one gallery photo unless you have been told otherwise. Listings on the", None),
    ("free tier display one gallery image, so anything beyond that is stored and never seen.", None),
    ("", None),
    ("Photographs must be ones we are allowed to publish. Send the credit and where each", None),
    ("came from with the folder - a photo we cannot show is worse than no photo.", None),
    ("", None),
    ("What happens to what you write", "head"),
    ("Nothing appears on the site automatically. Every listing arrives as a draft and somebody", None),
    ("on the team reads it before it is published.", None),
]

row = 1
for text, kind in GUIDE:
    cell = guide.cell(row=row, column=1, value=text)
    if kind == "title":
        cell.font = Font(size=16, bold=True)
    elif kind == "head":
        cell.font = Font(size=12, bold=True)
    row += 1

# --- the sheet they fill ---------------------------------------------------
sheet = book.create_sheet("Listings")

# --- the backing sheet for the long list -----------------------------------
#
# Excel caps an inline dropdown list at 255 characters, and sixty towns is
# roughly twice that. A list that long has to live in cells and be referenced by
# range, so it gets its own sheet - hidden, because it is machinery rather than
# something to fill in.
lists = book.create_sheet("Lists")
lists.column_dimensions["A"].width = 30
lists.cell(row=1, column=1, value="Towns already in the directory").font = Font(bold=True)

for offset, town in enumerate(TOWNS, start=2):
    lists.cell(row=offset, column=1, value=town)

TOWN_RANGE = f"Lists!$A$2:$A${len(TOWNS) + 1}"
lists.sheet_state = "hidden"

for index, (header, width, help_text, options, required, strict) in enumerate(COLUMNS, start=1):
    letter = get_column_letter(index)
    sheet.column_dimensions[letter].width = width

    head = sheet.cell(row=1, column=index, value=header)
    head.font = Font(bold=True, color="FFFFFF")
    head.fill = HEADER_FILL
    head.alignment = Alignment(vertical="center", wrap_text=True)

    hint = sheet.cell(row=2, column=index, value=help_text)
    hint.fill = REQUIRED_FILL if required else HELP_FILL
    hint.font = Font(size=9, italic=True, color="374151")
    hint.alignment = Alignment(vertical="top", wrap_text=True)

    rule = None

    if options is TOWNS:
        rule = DataValidation(type="list", formula1=TOWN_RANGE, allow_blank=True)
    elif options:
        rule = DataValidation(
            type="list", formula1='"' + ",".join(options) + '"', allow_blank=not required
        )
    elif header == "Rating / 5":
        # A number, not a list. Typing "four and a half" here reaches the
        # importer as nothing at all, which looks like a rating nobody checked.
        rule = DataValidation(type="decimal", operator="between", formula1=0, formula2=5)
        rule.error = "A Google rating is a number between 0 and 5, for example 4.6."
        rule.errorTitle = "Not a rating"

    if rule is None:
        continue

    if rule.error is None:
        rule.error = (
            "Pick one of the listed options - anything else is rejected by the importer."
            if strict
            else "That is not on the list. Continue only if it is genuinely a new one."
        )
        rule.errorTitle = "Not an accepted value" if strict else "Not on the list"

    # A locked list on Category is right, because a value off the list cannot be
    # imported at all. A locked list on Location would make the sheet unusable
    # the first time a village nobody has entered before comes up.
    rule.errorStyle = "stop" if strict else "warning"
    rule.showErrorMessage = True

    sheet.add_data_validation(rule)
    rule.add(f"{letter}3:{letter}500")

sheet.row_dimensions[1].height = 28
sheet.row_dimensions[2].height = 58
sheet.freeze_panes = "A3"

# One filled row, so the shape is obvious rather than described.
EXAMPLE = [
    "Beit el Qamar",
    "Guest Houses",
    "Keserwan District",
    "Faraya",
    "Guest house, mountain views, breakfast included",
    "",
    "$50-149",
    "4.6",
    "Year-round",
    "A restored stone house above Faraya with six rooms and a terrace looking over the valley.",
]

for index, value in enumerate(EXAMPLE, start=1):
    cell = sheet.cell(row=3, column=index, value=value)
    cell.font = Font(color="9CA3AF", italic=True)
    cell.alignment = Alignment(vertical="top", wrap_text=True)

sheet.cell(row=4, column=1, value="^ example, delete this row").font = Font(
    size=9, italic=True, color="9CA3AF"
)

book.save(OUT)
print(f"wrote {OUT}")
dropdowns = [c[0] for c in COLUMNS if c[3]]
print(f"  {len(COLUMNS)} columns, {len(TOWNS)} towns")
print("  dropdowns: " + ", ".join(dropdowns))
